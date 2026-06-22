"""Branded Excel workbook for a ReportModel of trade segments.

A combined Summary sheet (per-segment KPIs + overview), then each segment's
sheets: the drill sheet (sheet-metal size aggregation, or plumbing line-item
detail) plus its flat breakdowns. Tabs are trade-prefixed when >1 segment.
Excel keeps Arial and the red accent; the report styling is self-contained.
"""
from __future__ import annotations

import os
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import aggregate, brand, views
from .model import column_header, fmt_value


def _hx(c):
    return c.lstrip("#").upper()


WHITE, RED, BLACK = _hx(brand.WHITE), _hx(brand.FLOOR_FILL), _hx(brand.INK)
SUB, ALT, LINE, MID = _hx(brand.SUBTOTAL_FILL), _hx(brand.ROW_ALT), _hx(brand.ROW_LINE), _hx(brand.MID_GRAY)
F = brand.EXCEL_FONT
_thin = Side(style="thin", color=LINE)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fill(ws, row, c0, c1, hexc):
    pf = PatternFill(fill_type="solid", fgColor=hexc)
    for c in range(c0, c1 + 1):
        ws.cell(row=row, column=c).fill = pf


# A text cell whose value begins with one of these is treated as a formula by
# Excel (openpyxl even stores a leading "=" as a live formula). Labels here come
# from the untrusted raw export (system, drawing, spec, item names, bid name), so
# neutralize the leading character to keep the workbook inert. Applied only to
# text labels — numeric measure cells are written separately with a number format.
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r", "\n")


def _safe_text(value):
    if isinstance(value, str) and value[:1] in _FORMULA_LEAD:
        return "'" + value
    return value


def _label(ws, row, text, *, indent=0, bold=False, color="000000", size=9):
    cell = ws.cell(row=row, column=1, value=_safe_text(text))
    cell.font = Font(name=F, size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="left", indent=indent, vertical="center")
    cell.border = BORDER
    return cell


def _values(ws, row, columns, values, grand, *, row_type="group", bold=False, color="000000"):
    """Write the measure columns (col 2..) from a values dict.

    row_type 'group' fills group_only cols (and percentages); 'line' fills
    line_only cols. Columns flagged for the other type are left blank.
    """
    vals = aggregate.add_percentages(values, grand) if row_type == "group" else dict(values)
    for i, col in enumerate(columns):
        cell = ws.cell(row=row, column=2 + i)
        cell.font = Font(name=F, size=9, bold=bold, color=color)
        cell.alignment = Alignment(horizontal="right")
        cell.border = BORDER
        skip = (row_type == "line" and col.get("group_only")) or \
               (row_type == "group" and col.get("line_only"))
        if not skip:
            v = vals.get(col["key"])
            if v is not None:
                cell.value = v
                cell.number_format = col["excel"]


def _header(ws, row, columns, unit_labels, label_text="Section"):
    _label(ws, row, label_text, bold=True, color=WHITE)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for i, col in enumerate(columns):
        cell = ws.cell(row=row, column=2 + i, value=column_header(col, unit_labels))
        cell.font = Font(name=F, size=9, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        cell.border = BORDER
    _fill(ws, row, 1, 1 + len(columns), RED)
    ws.row_dimensions[row].height = 26


def _widths(ws, columns, label_w=34):
    ws.column_dimensions["A"].width = label_w
    for i, col in enumerate(columns):
        ws.column_dimensions[get_column_letter(2 + i)].width = 8 if col["kind"] == "pct" else 12


def _title_block(ws, ncols, title, subtitle):
    ws.cell(row=1, column=1, value=_safe_text(title)).font = Font(name=F, size=15, bold=True, color=BLACK)
    ws.cell(row=2, column=1, value=subtitle).font = Font(name=F, size=9, color=MID)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=ncols + 1)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=ncols + 1)


def _print_setup(ws, header_row):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"


def _generated():
    return date.today().strftime("%b %d, %Y")


# --------------------------------------------------------------- KPI cards
def _kpi_cards(seg):
    ul = seg["unit_labels"]
    k = seg["data"]["kpis"]
    cards = []
    for spec in seg["profile"].kpis:
        label = spec["label"].format(weight=ul.get("weight", ""), area=ul.get("area", ""),
                                     per_area=ul.get("per_area", ""))
        cards.append((fmt_value(spec["fmt"], k.get(spec["key"], 0.0)), label))
    return cards


# ----------------------------------------------------------------- sheets
def _summary_sheet(ws, model):
    ncols = 6
    title = model["title"] or "Trimble Take-off"
    _title_block(ws, ncols, title, "Take-off Summary  -  generated " + _generated())
    row = 4
    for seg in model["segments"]:
        ul = seg["unit_labels"]
        _label(ws, row, f"{seg['label']}  ({ul['display']})", bold=True, color=WHITE, size=11)
        _fill(ws, row, 1, ncols + 1, BLACK)
        row += 1
        cards = _kpi_cards(seg)
        for i, (val, lbl) in enumerate(cards):
            ws.cell(row=row, column=1 + i, value=val).font = Font(name=F, size=12, bold=True, color=_hx(brand.ACCENT))
            ws.cell(row=row + 1, column=1 + i, value=lbl).font = Font(name=F, size=8, color=MID)
        srcs = [s for s in seg["sources"]]
        note = f"{seg['data']['kpis']['line_count']:,} line items"
        if len(set(srcs)) > 1:
            note += "  from: " + ", ".join(sorted(set(srcs)))
        ws.cell(row=row + 2, column=1, value=note).font = Font(name=F, size=8, italic=True, color=MID)
        row += 4

        # Overview = top-level hierarchy nodes (floors for SM, systems for PL).
        cols = seg["profile"].summary_columns
        dim0 = seg["data"]["hierarchy_dims"][0]
        kind0 = {"floor": "floor", "system": "system"}.get(dim0, dim0)
        _header(ws, row, cols, ul, label_text=dim0.replace("_norm", "").title())
        row += 1
        grand = seg["data"]["grand"]
        for n in seg["data"]["hierarchy"]:
            _label(ws, row, views.label(n["key"], kind0))
            _values(ws, row, cols, n["totals"], grand)
            if row % 2 == 0:
                _fill(ws, row, 1, 1 + len(cols), ALT)
            row += 1
        _label(ws, row, "Total", bold=True)
        _values(ws, row, cols, grand, grand, bold=True, color=WHITE)
        _fill(ws, row, 1, 1 + len(cols), BLACK)
        row += 3
    ws.freeze_panes = "A4"
    _widths(ws, model["segments"][0]["profile"].summary_columns)


def _flat_sheet(ws, seg, flat, prefix):
    cols = seg["profile"].summary_columns
    ul = seg["unit_labels"]
    _widths(ws, cols)
    _title_block(ws, len(cols), prefix + flat["title"],
                 ul["display"] + "  -  generated " + _generated())
    hdr = 4
    _header(ws, hdr, cols, ul, label_text=flat["title"].replace("By ", ""))
    row = hdr + 1
    grand = seg["data"]["grand"]
    for n in flat["nodes"]:
        _label(ws, row, views.label(n["key"], flat["kind"]))
        _values(ws, row, cols, n["totals"], grand)
        if row % 2 == 0:
            _fill(ws, row, 1, 1 + len(cols), ALT)
        row += 1
    _label(ws, row, "TOTAL", bold=True, color=WHITE)
    _values(ws, row, cols, grand, grand, bold=True, color=WHITE)
    _fill(ws, row, 1, 1 + len(cols), BLACK)
    ws.freeze_panes = "A5"
    _print_setup(ws, hdr)


def _drill_size_agg(ws, seg, prefix):
    """Sheet-metal drill: Floor band -> Size rows (collapsible) -> System
    subtotal -> Floor total -> grand total."""
    cols = seg["profile"].summary_columns
    ul = seg["unit_labels"]
    _widths(ws, cols)
    _title_block(ws, len(cols), prefix + "By Floor / System / Size",
                 ul["display"] + "  -  generated " + _generated())
    ws.sheet_properties.outlinePr.summaryBelow = True
    hdr = 4
    _header(ws, hdr, cols, ul)
    row = hdr + 1
    grand = seg["data"]["grand"]
    for floor in seg["data"]["hierarchy"]:
        _label(ws, row, views.label(floor["key"], "floor"), bold=True, color=WHITE, size=10)
        _fill(ws, row, 1, 1 + len(cols), RED)
        row += 1
        for system in floor["children"]:
            for sz in system["children"] or []:
                _label(ws, row, views.label(sz["key"], "size"), indent=2)
                _values(ws, row, cols, sz["totals"], grand)
                if row % 2 == 0:
                    _fill(ws, row, 1, 1 + len(cols), ALT)
                ws.row_dimensions[row].outline_level = 1
                row += 1
            _label(ws, row, views.label(system["key"], "system") + " subtotal", indent=1, bold=True)
            _values(ws, row, cols, system["totals"], grand, bold=True)
            _fill(ws, row, 1, 1 + len(cols), SUB)
            row += 1
        _label(ws, row, views.label(floor["key"], "floor") + " total", bold=True)
        _values(ws, row, cols, floor["totals"], grand, bold=True)
        _fill(ws, row, 1, 1 + len(cols), SUB)
        row += 1
    _label(ws, row, "PROJECT TOTAL", bold=True, color=WHITE, size=10)
    _values(ws, row, cols, grand, grand, bold=True, color=WHITE)
    _fill(ws, row, 1, 1 + len(cols), BLACK)
    ws.freeze_panes = "A5"
    _print_setup(ws, hdr)


def _drill_line_items(ws, seg, prefix):
    """Plumbing drill: System (overview row) -> Material Spec subtotal ->
    line items (collapsible), with per-line pricing detail. When the export
    carries Floor data, views.build prepends a 'floor' level, so a Floor band
    wraps each System group (the level shift is handled by reading the dim
    order, not a fixed depth)."""
    cols = seg["profile"].detail_columns
    ul = seg["unit_labels"]
    _widths(ws, cols, label_w=42)
    _title_block(ws, len(cols), prefix + "By System / Spec  (line detail)",
                 ul["display"] + "  -  generated " + _generated())
    ws.sheet_properties.outlinePr.summaryBelow = False
    hdr = 4
    _header(ws, hdr, cols, ul)
    grand = seg["data"]["grand"]
    dims = seg["data"]["hierarchy_dims"]
    floored = dims and dims[0] == "floor"
    row = [hdr + 1]   # boxed so the nested writers can advance it

    def emit_system(system, base_outline):
        # System overview row (carries the summary measures).
        _label(ws, row[0], views.label(system["key"], "system"), bold=True, color=WHITE, size=10)
        _values(ws, row[0], cols, system["totals"], grand, bold=True, color=WHITE, row_type="group")
        _fill(ws, row[0], 1, 1 + len(cols), RED)
        if base_outline:
            ws.row_dimensions[row[0]].outline_level = base_outline
        row[0] += 1
        for spec in system["children"] or []:
            _label(ws, row[0], views.label(spec["key"], "spec"), indent=1, bold=True)
            _values(ws, row[0], cols, spec["totals"], grand, bold=True, row_type="group")
            _fill(ws, row[0], 1, 1 + len(cols), SUB)
            ws.row_dimensions[row[0]].outline_level = base_outline + 1
            row[0] += 1
            lines = sorted(spec["records"], key=lambda r: (str(r.get("sorted_size") or ""),
                                                           str(r.get("item_name") or "")))
            for rec in lines:
                name = str(rec.get("item_name") or "")
                sz = str(rec.get("size") or "")
                _label(ws, row[0], f"{name}  {sz}".strip(), indent=2)
                _values(ws, row[0], cols, rec, grand, row_type="line")
                if row[0] % 2 == 0:
                    _fill(ws, row[0], 1, 1 + len(cols), ALT)
                ws.row_dimensions[row[0]].outline_level = base_outline + 2
                row[0] += 1

    if floored:
        for floor in seg["data"]["hierarchy"]:
            _label(ws, row[0], views.label(floor["key"], "floor"), bold=True, color=WHITE, size=10)
            _fill(ws, row[0], 1, 1 + len(cols), BLACK)
            row[0] += 1
            for system in floor["children"] or []:
                emit_system(system, base_outline=1)
            _label(ws, row[0], views.label(floor["key"], "floor") + " total", bold=True)
            _values(ws, row[0], cols, floor["totals"], grand, bold=True, row_type="group")
            _fill(ws, row[0], 1, 1 + len(cols), SUB)
            row[0] += 1
    else:
        for system in seg["data"]["hierarchy"]:
            emit_system(system, base_outline=0)

    _label(ws, row[0], "PROJECT TOTAL", bold=True, color=WHITE, size=10)
    _values(ws, row[0], cols, grand, grand, bold=True, color=WHITE, row_type="group")
    _fill(ws, row[0], 1, 1 + len(cols), BLACK)
    ws.freeze_panes = "A5"
    _print_setup(ws, hdr)


def write(model: dict, out_path) -> Path:
    out_path = Path(out_path)
    wb = Workbook()
    _summary_sheet(wb.active, model)
    wb.active.title = "Summary"

    multi = model["multi"]
    for seg in model["segments"]:
        prefix = (seg["abbr"] + " ") if multi else ""
        drill_tab = (prefix + ("Detail" if seg["profile"].detail_mode == "line_items"
                               else "Floor + System"))[:31]
        ws = wb.create_sheet(drill_tab)
        if seg["profile"].detail_mode == "line_items":
            _drill_line_items(ws, seg, prefix)
        else:
            _drill_size_agg(ws, seg, prefix)
        for flat in seg["data"]["flats"]:
            tab = (prefix + flat["tab"])[:31]
            _flat_sheet(wb.create_sheet(tab), seg, flat, prefix)

    return _save_lock_safe(wb, out_path)


def _save_lock_safe(wb, out_path: Path) -> Path:
    tmp = out_path.with_name(f"{out_path.stem}.tmp_{os.getpid()}.xlsx")
    wb.save(str(tmp))
    try:
        os.replace(tmp, out_path)
        return out_path
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d-%H%M")
        alt = out_path.with_name(f"{out_path.stem} (locked-{stamp}).xlsx")
        os.replace(tmp, alt)
        print(f"WARNING: '{out_path.name}' is open. Wrote '{alt.name}' instead.")
        return alt
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
