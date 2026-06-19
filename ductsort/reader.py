"""Read a Trimble AutoBid raw export (sheet metal or plumbing) into records.

Locates the raw-data sheet by its header columns rather than name or position,
so it handles both the simple layout and the "Pivot" layout where Raw Data sits
far right among many extra tabs. The trade (sheet metal vs plumbing) is detected
from the columns; metadata comes from the QPSummary defined names.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from . import model, trades


class RawDataError(Exception):
    """Raised when the file does not look like an AutoBid raw export."""


_PLACEHOLDERS = ("will have", "filled in by autobid")
_MAX_SCAN = 15


def _sample(ws, n=_MAX_SCAN):
    out = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=n, values_only=True)):
        out.append(row)
    return out


def _header_in(rows):
    """Return (row_index, headers_list) for the first row carrying the common
    signature, else (None, None)."""
    sig = set(trades.COMMON_SIGNATURE)
    for ri, row in enumerate(rows):
        names = {str(v).strip() for v in row if v is not None}
        if sig.issubset(names):
            return ri, list(row)
    return None, None


def _find_raw_sheet(wb):
    """Pick the raw-data worksheet across all sheets. Returns (ws, hdr_idx,
    headers) or None. Prefers name hints, then the richest header match."""
    candidates = []
    for ws in wb.worksheets:
        rows = _sample(ws)
        hdr_idx, headers = _header_in(rows)
        if hdr_idx is None:
            continue
        names = {str(h).strip() for h in headers if h is not None}
        trade = trades.detect_trade(headers)
        name_hit = any(h in ws.title.lower() for h in ("raw data", "rawdata"))
        score = (1 if trade else 0, 1 if name_hit else 0, len(names))
        candidates.append((score, ws, hdr_idx, headers))
    if not candidates:
        return None
    candidates.sort(key=lambda c: c[0], reverse=True)
    _, ws, hdr_idx, headers = candidates[0]
    return ws, hdr_idx, headers


def _defined_value(wb, name):
    """Resolve a defined name to its first cell value, robust across openpyxl
    versions; returns None on any failure."""
    try:
        dn = wb.defined_names[name]
    except Exception:
        dn = None
        try:
            for d in wb.defined_names.definedName:   # older openpyxl
                if d.name == name:
                    dn = d
                    break
        except Exception:
            dn = None
    if dn is None:
        return None
    try:
        for sheet, coord in dn.destinations:
            return wb[sheet][coord].value
    except Exception:
        return None
    return None


_LABELS = {
    "bid name:": "bid_name", "project:": "bid_name",
    "bid number:": "bid_number", "bid i.d.:": "bid_number", "bid id:": "bid_number",
    "company name:": "company", "filter:": "filter",
}


def _read_metadata(wb) -> dict:
    """Bid name / number / company / filter, via defined names then label scan."""
    meta = {"bid_name": "", "bid_number": "", "company": "", "filter": ""}
    for qp, key in (("QPSummaryBid_Name", "bid_name"),
                    ("QPSummaryBid_Number", "bid_number"),
                    ("QPSummaryCO_Name", "company"),
                    ("QPSummaryFilter", "filter")):
        v = _defined_value(wb, qp)
        if v and not any(p in str(v).lower() for p in _PLACEHOLDERS):
            meta[key] = str(v).strip()
    if "Summary" in wb.sheetnames and not (meta["bid_name"] and meta["bid_number"]):
        for row in wb["Summary"].iter_rows(values_only=True):
            for i, v in enumerate(row):
                lab = _LABELS.get(str(v).strip().lower()) if isinstance(v, str) else None
                if lab and not meta[lab]:
                    nxt = row[i + 1] if i + 1 < len(row) else None
                    if nxt and not any(p in str(nxt).lower() for p in _PLACEHOLDERS):
                        meta[lab] = str(nxt).strip()
    return meta


_BLANK_DEFAULTS = {
    "floor": "(no floor)", "system": "(no system)", "drawing": "(no drawing)",
    "specs": "(no spec)", "material_spec": "(no spec)", "trade": "(no trade)",
    "item_type": "(no item type)",
}


def load(path: str | Path):
    """Load a raw export. Returns (records, metadata, profile).

    Each record carries cleaned dimensions, a normalised `size_norm`, and the
    trade's derived measures (material_cost, labour_hrs, weight, ...).
    """
    path = Path(path)
    if not path.exists():
        raise RawDataError(f"File not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        found = _find_raw_sheet(wb)
        if not found:
            raise RawDataError(
                "Could not find the take-off data. Expected a sheet with a header "
                "row containing: " + ", ".join(trades.COMMON_SIGNATURE) + ".")
        ws, hdr_idx, headers = found
        profile = trades.detect_trade(headers)
        if profile is None:
            raise RawDataError(
                "Found a data sheet but could not tell whether it is sheet metal "
                "or plumbing (missing the expected cost/measure columns).")

        col_of = {}
        for ci, h in enumerate(headers):
            key = profile.column_map.get(str(h).strip() if h is not None else "")
            if key and key not in col_of:
                col_of[key] = ci

        tdims = trades.text_dims(profile)
        records = []
        for row in ws.iter_rows(min_row=hdr_idx + 2, values_only=True):
            rec = {}
            empty = True
            for key, ci in col_of.items():
                v = row[ci] if ci < len(row) else None
                if v is not None and not (isinstance(v, str) and not v.strip()):
                    empty = False
                rec[key] = v
            if empty:
                continue
            for k in tdims:
                val = rec.get(k)
                rec[k] = str(val).strip() if val is not None else ""
            for k, default in _BLANK_DEFAULTS.items():
                if k in rec and not rec[k]:
                    rec[k] = default
            rec["size_norm"] = model.normalize_size(rec.get("size"))
            profile.derive(rec)
            records.append(rec)

        if not records:
            raise RawDataError("The take-off sheet has a header but no data rows.")
        return records, _read_metadata(wb), profile
    finally:
        wb.close()
